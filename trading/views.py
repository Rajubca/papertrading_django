import json
import io
import os
import glob
import requests
import pandas as pd
from decimal import Decimal, InvalidOperation
from datetime import timedelta

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.contrib.auth import login
from django.contrib.auth.models import User
from django.db.models import Q
from django.contrib.auth.views import LoginView, LogoutView, PasswordResetView, PasswordResetDoneView, PasswordResetConfirmView, PasswordResetCompleteView

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Stock, Portfolio, Transaction, Holding, PortfolioReport, HoldingReport, Watchlist, Profile, NSEData
from .forms import TradeForm, PortfolioForm, WatchlistForm, UserRegisterForm, UserUpdateForm, ProfileUpdateForm

# Helper function to handle trade logic
def update_portfolio_after_trade(portfolio, stock, quantity, price, transaction_type):
    # Fetch existing holding or create dummy if None
    holding = Holding.objects.filter(portfolio=portfolio, stock=stock).first()

    current_qty = holding.quantity if holding else 0
    avg_price = holding.average_buy_price if holding else Decimal('0.00')

    trade_value = Decimal(quantity) * Decimal(price)

    # Update Cash Balance
    if transaction_type == 'BUY':
        portfolio.cash_balance -= trade_value
    else: # SELL
        portfolio.cash_balance += trade_value

    # Calculate New Holding State
    new_qty = current_qty
    new_avg_price = avg_price

    if transaction_type == 'BUY':
        # Buying shares
        if current_qty >= 0:
            # Case 1: Increasing Long Position (or opening)
            total_cost = (Decimal(current_qty) * avg_price) + trade_value
            new_qty = current_qty + quantity
            if new_qty > 0:
                new_avg_price = total_cost / Decimal(new_qty)
            else:
                new_avg_price = Decimal('0.00')
        else:
            # Case 2: Covering Short Position
            # Buying back shares you owe.
            if abs(current_qty) >= quantity:
                # 2a: Partial or Full Cover (remaining is still short or flat)
                new_qty = current_qty + quantity
                # Avg Price (Avg Sell Price) remains SAME for the remaining short position.
                if new_qty == 0:
                    new_avg_price = Decimal('0.00')
            else:
                # 2b: Flip from Short to Long
                remaining_short = abs(current_qty)
                excess_buy = quantity - remaining_short

                # Close the short part (P/L realized implicitly by cash change vs initial short proceeds)
                # Open new Long position
                new_qty = excess_buy
                new_avg_price = Decimal(price)

    else: # SELL
        # Selling shares
        if current_qty > 0:
            # Case 3: Decreasing Long Position (or closing)
            if current_qty >= quantity:
                # 3a: Partial or Full Sell
                new_qty = current_qty - quantity
                # Avg Price (Avg Buy Price) remains SAME for remaining long position.
                if new_qty == 0:
                    new_avg_price = Decimal('0.00')
            else:
                # 3b: Flip from Long to Short
                remaining_long = current_qty
                excess_sell = quantity - remaining_long

                # Close the long part
                # Open new Short position
                new_qty = -excess_sell
                new_avg_price = Decimal(price)
        else:
            # Case 4: Increasing Short Position (or opening)
            # Adding to short position
            # Weighted Average of SELL prices
            total_value_short = (Decimal(abs(current_qty)) * avg_price) + trade_value
            new_qty = current_qty - quantity
            if new_qty != 0:
                new_avg_price = total_value_short / Decimal(abs(new_qty))
            else:
                new_avg_price = Decimal('0.00')

    # Save or Delete Holding
    if new_qty == 0:
        if holding:
            holding.delete()
            holding = None
    else:
        if not holding:
            holding = Holding(portfolio=portfolio, stock=stock, quantity=0, average_buy_price=0)
        holding.quantity = new_qty
        holding.average_buy_price = new_avg_price
        holding.save()

    portfolio.save()
    return holding

@login_required
def dashboard(request):
    portfolio_id = request.GET.get('portfolio') or request.session.get('active_portfolio_id')
    portfolios = Portfolio.objects.filter(user=request.user, visibility='PUBLIC')

    if not portfolios.exists():
        template_name = 'trading/partials/dashboard_content.html' if request.headers.get('HX-Request') == 'true' else 'trading/dashboard.html'
        return render(request, template_name, {
            'portfolios': None, 'holdings': [], 'transactions': [], 'performance_data': json.dumps([])
        })

    if portfolio_id:
        try:
            portfolio = portfolios.get(id=portfolio_id)
        except Portfolio.DoesNotExist:
            portfolio = portfolios.first()
    else:
        portfolio = portfolios.first()

    request.session['active_portfolio_id'] = portfolio.id
    holdings = Holding.objects.filter(portfolio=portfolio)
    transactions = Transaction.objects.filter(portfolio=portfolio).order_by('-timestamp')[:10]
    performance_data = generate_performance_data(portfolio)

    total_value = portfolio.total_value # Use property
    initial_cash = getattr(portfolio, 'initial_cash', portfolio.cash_balance) # Fallback logic

    # Calculate total P/L
    profit_loss = total_value - initial_cash
    profit_loss_percentage = Decimal('0.00')
    if initial_cash:
        profit_loss_percentage = (profit_loss / initial_cash * Decimal('100')).quantize(Decimal('0.01'))

    context = {
        'portfolio': portfolio,
        'portfolios': portfolios,
        'holdings': holdings,
        'transactions': transactions,
        'performance_data': json.dumps(performance_data),
        'total_value': total_value,
        'profit_loss': profit_loss,
        'profit_loss_percentage': profit_loss_percentage,
        'initial_cash': initial_cash,
    }

    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/dashboard_content.html', context)
    return render(request, 'trading/dashboard.html', context)

def generate_performance_data(portfolio):
    total_value = float(portfolio.total_value)
    performance_data = []
    today = timezone.now().date()
    for i in range(30, 0, -1):
        date = today - timedelta(days=i)
        value = total_value * (0.95 + 0.1 * (i / 30))
        performance_data.append({"date": date.isoformat(), "value": round(value, 2)})
    return performance_data

def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('trading:dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

@login_required
def stock_list(request):
    stocks = Stock.objects.all()
    query = request.GET.get('q')
    if query:
        stocks = stocks.filter(Q(symbol__icontains=query) | Q(name__icontains=query))

    context = {'stocks': stocks, 'query': query}
    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/stock_list_content.html', context)
    return render(request, 'trading/stock_list.html', context)

def stock_search(request):
    query = request.GET.get('q', '')
    if query:
        stocks = Stock.objects.filter(Q(symbol__icontains=query) | Q(name__icontains=query))[:10]
    else:
        stocks = Stock.objects.none()
    results = [{'id': s.id, 'symbol': s.symbol, 'name': s.name, 'current_price': str(s.current_price)} for s in stocks]
    return JsonResponse(results, safe=False)

@login_required
def stock_detail(request, pk):
    stock = get_object_or_404(Stock, pk=pk)
    user_holdings = Holding.objects.filter(portfolio__user=request.user, stock=stock)
    if request.user.is_staff or request.user.is_superuser:
        recent_transactions = Transaction.objects.filter(stock=stock).order_by('-timestamp')[:10]
    else:
        recent_transactions = Transaction.objects.filter(portfolio__user=request.user, stock=stock).order_by('-timestamp')[:10]
    return render(request, 'trading/stock_detail.html', {'stock': stock, 'user_holdings': user_holdings, 'recent_transactions': recent_transactions})

@login_required
def trade_stock(http_request, stock_id=None):
    portfolios = Portfolio.objects.filter(user=http_request.user, visibility='PUBLIC')
    if not portfolios.exists():
        messages.error(http_request, "You need to create a portfolio first!")
        if http_request.headers.get('HX-Request') == 'true':
             return HttpResponse(status=204, headers={'HX-Redirect': '/portfolios/create/'})
        return redirect('trading:create_portfolio')

    portfolio_id = http_request.session.get('active_portfolio_id')
    portfolio = None
    if portfolio_id:
        try:
            portfolio = portfolios.get(id=portfolio_id)
        except Portfolio.DoesNotExist:
            portfolio = portfolios.first()
    else:
        portfolio = portfolios.first()

    stocks = Stock.objects.all()
    initial_data = {}
    selected_stock = None
    if stock_id:
        selected_stock = get_object_or_404(Stock, id=stock_id)
        initial_data['stock'] = selected_stock
        initial_data['price_per_share'] = float(selected_stock.current_price)

    if http_request.method == 'POST':
        form = TradeForm(http_request.POST, initial=initial_data, user=http_request.user)
        if form.is_valid():
            transaction = form.save(commit=False)
            selected_portfolio = form.cleaned_data['portfolio']
            transaction.portfolio = selected_portfolio
            transaction.user = http_request.user

            # Using helper to update logic
            update_portfolio_after_trade(
                selected_portfolio,
                transaction.stock,
                transaction.quantity,
                transaction.price_per_share,
                transaction.transaction_type
            )

            transaction.save()

            # Generate Report Automatically
            try:
                selected_portfolio.generate_report()
            except Exception as e:
                print(f"Failed to generate report: {e}")

            http_request.session['active_portfolio_id'] = selected_portfolio.id
            messages.success(http_request, f"Successfully {transaction.transaction_type.lower()}ed {transaction.quantity} shares of {transaction.stock.symbol}")

            if http_request.headers.get('HX-Request') == 'true':
                return HttpResponse(status=204, headers={'HX-Trigger': 'portfolioCreated'})

            return redirect('trading:dashboard')
    else:
        form = TradeForm(initial=initial_data, user=http_request.user)
        if portfolio:
            form.fields['portfolio'].initial = portfolio

    if not selected_stock and stocks.exists():
        selected_stock = stocks.first()

    context = {
        'form': form,
        'portfolio': portfolio,
        'portfolios': portfolios,
        'stocks': stocks,
        'selected_stock': selected_stock,
        'stocks_count': stocks.count(),
        'portfolio_cash_balance': float(portfolio.cash_balance) if portfolio else 0
    }

    if http_request.headers.get('HX-Request') == 'true':
        return render(http_request, 'trading/partials/modal_trade_stock.html', context)

    return render(http_request, 'trading/trade.html', context)

@csrf_exempt
@login_required
@require_POST
def calculate_order_summary(http_request):
    try:
        if http_request.content_type == 'application/json':
            data = json.loads(http_request.body)
            stock_id = data.get("stock_id")
            quantity = data.get("quantity", 0)
            price = data.get("price", 0)
            transaction_type = data.get("transaction_type", "BUY")
            portfolio_id = data.get("portfolio_id")
        else:
            stock_id = http_request.POST.get("stock_id")
            quantity = http_request.POST.get("quantity", 0)
            price = http_request.POST.get("price", 0)
            transaction_type = http_request.POST.get("transaction_type", "BUY")
            portfolio_id = http_request.POST.get("portfolio_id")

        stock_id = int(stock_id) if stock_id else None
        quantity = int(quantity) if quantity else 0
        price = Decimal(str(price)) if price else Decimal('0')

        if not stock_id:
            return JsonResponse({"success": False, "error": "Stock ID is required"}, status=400)

        stock = get_object_or_404(Stock, id=stock_id)

        if portfolio_id:
            portfolio = get_object_or_404(Portfolio, id=portfolio_id, user=http_request.user, visibility='PUBLIC')
        else:
            portfolio = Portfolio.objects.filter(user=http_request.user, visibility='PUBLIC').first()
            if not portfolio:
                return JsonResponse({"success": False, "error": "No portfolio found"}, status=400)

        total_amount = Decimal(quantity) * price

        # New Logic for Calc
        cash_after = portfolio.cash_balance
        if transaction_type == 'BUY':
            cash_after -= total_amount
        else:
            cash_after += total_amount

        # Check existing shares
        try:
            holding = Holding.objects.get(portfolio=portfolio, stock=stock)
            owned_shares = holding.quantity
        except Holding.DoesNotExist:
            owned_shares = 0

        response_data = {
            "success": True,
            "quantity": quantity,
            "price": float(price),
            "total": float(total_amount),
            "cash_after": float(cash_after),
            "current_price": float(stock.current_price),
            "owned_shares": owned_shares,
            "portfolio_cash": float(portfolio.cash_balance)
        }
        return JsonResponse(response_data)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

@login_required
def transaction_list(request):
    portfolio_id = request.GET.get('portfolio') or request.session.get('active_portfolio_id')
    portfolios = Portfolio.objects.filter(user=request.user, visibility='PUBLIC')

    if not portfolios.exists():
        context = {'transactions': [], 'portfolios': None}
        if request.headers.get('HX-Request') == 'true':
            return render(request, 'trading/partials/transaction_list_content.html', context)
        return render(request, 'trading/transaction_list.html', context)

    if portfolio_id:
        try:
            portfolio = portfolios.get(id=portfolio_id)
        except Portfolio.DoesNotExist:
            portfolio = portfolios.first()
    else:
        portfolio = portfolios.first()

    transactions = portfolio.transaction_set.all().select_related('stock').order_by('-timestamp')
    context = {'transactions': transactions, 'portfolio': portfolio, 'portfolios': portfolios}

    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/transaction_list_content.html', context)
    return render(request, 'trading/transaction_list.html', context)

@login_required
def reports(request):
    user_portfolios = Portfolio.objects.filter(user=request.user)
    reports = PortfolioReport.objects.filter(portfolio__in=user_portfolios).order_by('-report_date', '-created_at')
    context = {'reports': reports}
    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/reports_content.html', context)
    return render(request, 'trading/reports.html', context)

@login_required
def watchlists(request):
    watchlists = request.user.watchlists.all().prefetch_related('stocks')
    form = WatchlistForm()
    if request.method == 'POST':
        form = WatchlistForm(request.POST)
        if form.is_valid():
            watchlist = form.save(commit=False)
            watchlist.user = request.user
            watchlist.save()
            form.save_m2m()
            messages.success(request, f"Watchlist '{watchlist.name}' created")
            return redirect('watchlists')
    context = {'watchlists': watchlists, 'form': form}
    if request.htmx:
        return render(request, 'trading/partials/watchlist_list.html', context)
    return render(request, 'trading/watchlists.html', context)

@login_required
def update_stock_price(request, pk):
    if request.method == 'POST' and request.htmx:
        stock = get_object_or_404(Stock, pk=pk)
        new_price = Decimal(request.POST.get('price'))
        stock.current_price = new_price # Direct update or method? Model had update_price method? No, just field.
        stock.save() # Assuming no method
        return render(request, 'trading/partials/stock_price.html', {'stock': stock})
    return JsonResponse({'error': 'Invalid request'}, status=400)

def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            profile = user.profile
            profile.phone = form.cleaned_data.get('phone')
            profile.birth_date = form.cleaned_data.get('birth_date')
            profile.save()
            messages.success(request, f'Account created for {user.username}!')
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'registration/register.html', {'form': form})

@staff_member_required
def manage_portfolio_visibility(request):
    if request.method == 'POST':
        portfolio_id = request.POST.get('portfolio_id')
        visibility = request.POST.get('visibility')
        if portfolio_id and visibility:
            portfolio = get_object_or_404(Portfolio, id=portfolio_id)
            portfolio.visibility = visibility
            portfolio.save()
            messages.success(request, f"Updated visibility for {portfolio.name}")
    portfolios = Portfolio.objects.all().select_related('user').order_by('user__username', 'name')
    return render(request, 'trading/admin_portfolio_visibility.html', {'portfolios': portfolios})

@login_required
def profile(request):
    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, instance=request.user.profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            messages.success(request, 'Your account has been updated!')
            return redirect('profile')
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=request.user.profile)
    return render(request, 'registration/profile.html', {'u_form': u_form, 'p_form': p_form})

def portfolio_list(request):
    portfolios = Portfolio.objects.filter(user=request.user) # All user portfolios
    context = {'portfolios': portfolios}
    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/portfolio_list_content.html', context)
    return render(request, 'trading/portfolio_list.html', context)

@login_required
def create_portfolio(request):
    form = PortfolioForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            portfolio = form.save(commit=False)
            portfolio.user = request.user
            portfolio.save()
            messages.success(request, f"Portfolio '{portfolio.name}' created successfully!")

            if request.headers.get('HX-Request') == 'true':
                return HttpResponse(status=204, headers={'HX-Trigger': 'portfolioCreated'})

            return redirect('trading:portfolio_list')

    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/modal_create_portfolio.html', {'form': form})

    return render(request, 'trading/create_portfolio.html', {'form': form})

@login_required
def delete_portfolio(request, pk):
    portfolio = get_object_or_404(Portfolio, pk=pk, user=request.user)
    if request.method == 'POST':
        portfolio.delete()
        messages.success(request, f"Portfolio '{portfolio.name}' deleted successfully.")
        return redirect('trading:portfolio_list')
    return render(request, 'trading/portfolio_confirm_delete.html', {'portfolio': portfolio})

@login_required
def report_detail(request, pk):
    report = get_object_or_404(PortfolioReport, pk=pk, portfolio__user=request.user)
    recent_transactions = Transaction.objects.filter(portfolio=report.portfolio).select_related('stock').order_by('-timestamp')[:10]

    if request.GET.get('export') == 'pdf':
        return generate_pdf_report(report)
    elif request.GET.get('export') == 'excel':
        return generate_excel_report(report)

    context = {'report': report, 'recent_transactions': recent_transactions}
    if request.headers.get('HX-Request') == 'true':
        return render(request, 'trading/partials/report_detail_content.html', context)
    return render(request, 'trading/report_detail.html', context)

@login_required
@require_GET
def stock_transactions_api(request, stock_id):
    try:
        stock = Stock.objects.get(id=stock_id)
        report_id = request.GET.get('report_id')
        if report_id:
            report = PortfolioReport.objects.get(id=report_id, portfolio__user=request.user)
            transactions = Transaction.objects.filter(portfolio=report.portfolio, stock=stock, timestamp__lte=report.report_date).order_by('-timestamp')[:20]
        else:
            transactions = Transaction.objects.filter(portfolio__user=request.user, stock=stock).order_by('-timestamp')[:20]

        data = [{'timestamp': t.timestamp.isoformat(), 'type': t.transaction_type, 'qty': t.quantity, 'price': float(t.price_per_share), 'total': float(t.total_cost)} for t in transactions]
        return JsonResponse({'success': True, 'stock_symbol': stock.symbol, 'transactions': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def generate_pdf_report(report):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"Portfolio Report - {report.report_date}", styles['Title']))

    data = [
        ['Total Value', f'{report.total_value:,.2f}'],
        ['Cash Balance', f'{report.cash_balance:,.2f}'],
        ['Invested Value', f'{report.investment_value:,.2f}']
    ]
    t = Table(data)
    t.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black)]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    if report.holding_reports.exists():
        elements.append(Paragraph("Holdings", styles['Heading2']))
        h_data = [['Symbol', 'Qty', 'Avg Price', 'Current', 'P/L %']]
        for hr in report.holding_reports.all():
            h_data.append([
                hr.holding.stock.symbol,
                str(hr.quantity),
                f'{hr.average_price:.2f}',
                f'{hr.current_price:.2f}',
                f'{hr.profit_loss_percentage:.2f}%'
            ])
        t2 = Table(h_data)
        t2.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 1, colors.black)]))
        elements.append(t2)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{report.id}.pdf"'
    return response

def generate_excel_report(report):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Report {report.report_date}"

    # Headers
    ws['A1'] = "Portfolio Report"
    ws['A1'].font = Font(size=14, bold=True)

    ws['A3'] = "Date"
    ws['B3'] = report.report_date

    ws['A4'] = "Total Value"
    ws['B4'] = report.total_value

    ws['A5'] = "Cash Balance"
    ws['B5'] = report.cash_balance

    ws['A6'] = "Invested Value"
    ws['B6'] = report.investment_value

    # Holdings
    ws['A8'] = "Holdings Breakdown"
    ws['A8'].font = Font(bold=True)

    headers = ['Symbol', 'Quantity', 'Avg Price', 'Current Price', 'Current Value', 'P/L', 'P/L %']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    row_num = 10
    for hr in report.holding_reports.all():
        ws.cell(row=row_num, column=1).value = hr.holding.stock.symbol
        ws.cell(row=row_num, column=2).value = hr.quantity
        ws.cell(row=row_num, column=3).value = hr.average_price
        ws.cell(row=row_num, column=4).value = hr.current_price
        ws.cell(row=row_num, column=5).value = hr.current_value
        ws.cell(row=row_num, column=6).value = hr.profit_loss
        ws.cell(row=row_num, column=7).value = hr.profit_loss_percentage
        row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="portfolio_report_{report.report_date}.xlsx"'
    return response

@login_required
def generate_report(request):
    portfolio = Portfolio.objects.filter(user=request.user).first()
    try:
        report = portfolio.generate_report()
        messages.success(request, f"Report generated for {report.report_date}")
        return redirect('trading:report_detail', pk=report.id)
    except Exception as e:
        messages.error(request, f"Error generating report: {str(e)}")
        return redirect('trading:reports')

@staff_member_required
def portfolio_manager(request):
    portfolios = Portfolio.objects.all()
    return render(request, 'trading/portfolio_manager.html', {'portfolios': portfolios})

@staff_member_required
def bulk_update_visibility(request):
    if request.method == 'POST':
        Portfolio.objects.filter(id__in=request.POST.getlist('portfolio_ids')).update(visibility=request.POST.get('visibility'))
    return redirect('trading:portfolio_manager')

@staff_member_required
def toggle_portfolio_visibility(request, portfolio_id):
    p = get_object_or_404(Portfolio, id=portfolio_id)
    p.visibility = 'PRIVATE' if p.visibility == 'PUBLIC' else 'PUBLIC'
    p.save()
    return redirect('trading:portfolio_manager')

@csrf_exempt
@require_POST
@login_required
def calculate_order(request):
    try:
        data = json.loads(request.body)
        stock_id = data.get('stock_id')
        quantity = int(data.get('quantity', 0))
        price = float(data.get('price', 0))
        transaction_type = data.get('transaction_type', 'BUY')

        stock = get_object_or_404(Stock, id=stock_id)
        portfolio = Portfolio.objects.filter(user=request.user).first()

        if not portfolio:
            return JsonResponse({'error': 'No portfolio found'}, status=400)

        total_cost = quantity * price

        # Check cash for BUY
        if transaction_type == 'BUY' and total_cost > portfolio.cash_balance:
             return JsonResponse({'error': f'Insufficient funds.'}, status=400)

        # Removed SELL restriction

        cash_after = portfolio.cash_balance - total_cost if transaction_type == 'BUY' else portfolio.cash_balance + total_cost

        return JsonResponse({
            'success': True,
            'total_cost': total_cost,
            'cash_after': cash_after,
            'current_cash': portfolio.cash_balance,
            'can_execute': True
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_stock_price_api(request, stock_id):
    try:
        stock = Stock.objects.get(id=stock_id)
        return JsonResponse({'success': True, 'price': float(stock.current_price), 'symbol': stock.symbol})
    except Stock.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Stock not found'}, status=404)

def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin, redirect_field_name=None, login_url='trading:dashboard')
def download_nse_data(request):
    downloaded_files = get_downloaded_files()

    if request.method == 'POST':
        index_name = request.POST.get('index_name')
        if not index_name:
            messages.error(request, "No index selected")
            return redirect('trading:data_download')

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"{index_name.replace(' ', '_')}_{timestamp}.csv"
        save_path = os.path.join(settings.BASE_DIR, 'data', 'nse', file_name)

        success = download_nse_csv(index_name, save_path)
        if success:
            messages.success(request, f"Successfully downloaded {index_name} data")
            try:
                # Corrected line to unpack all three values
                created_count, updated_count, skipped_count = process_nse_csv(save_path)
                messages.success(request,
                                 f"Updated stocks: {created_count} created, {updated_count} updated, {skipped_count} skipped due to invalid data")
            except Exception as e:
                messages.error(request, f"Failed to process CSV: {str(e)}")
        else:
            messages.error(request, f"Failed to download {index_name} data")

        return redirect('trading:data_download')

    context = {
        'downloaded_files': downloaded_files
    }
    return render(request, 'trading/data_download.html', context)


def download_nse_csv(index_name, save_path):
    url = (
        "https://www.nseindia.com/api/equity-stockIndices"
        f"?csv=true&index={index_name.replace(' ', '%20').replace('&', '%26')}"
        "&selectValFormat=crores"
    )
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.nseindia.com/market-data/live-equity-market",
    }
    try:
        with requests.Session() as session:
            session.get("https://www.nseindia.com/market-data/live-equity-market", headers=headers)
            response = session.get(url, headers=headers)
            response.raise_for_status()
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            with open(save_path, "wb") as f:
                f.write(response.content)
        return True
    except Exception as e:
        print(f"Failed to download {index_name}: {str(e)}")
        return False

def get_downloaded_files():
    data_dir = os.path.join(settings.BASE_DIR, 'data', 'nse')
    downloaded_files = []
    os.makedirs(data_dir, exist_ok=True)
    csv_files = glob.glob(os.path.join(data_dir, '*.csv'))
    for file_path in csv_files:
        file_name = os.path.basename(file_path)
        file_size = os.path.getsize(file_path)
        modified_time = os.path.getmtime(file_path)
        size_mb = file_size / (1024 * 1024)
        downloaded_files.append({
            'name': file_name,
            'size': f"{size_mb:.2f} MB",
            'path': file_path,
            'modified': modified_time
        })
    return downloaded_files

def view_file(request, filename):
    file_path = os.path.join(settings.BASE_DIR, 'data', 'nse', filename)
    if not os.path.exists(file_path):
        return HttpResponse("File not found", status=404)
    if not filename.endswith('.csv') or '..' in filename or '/' in filename:
        return HttpResponse("Invalid file", status=400)
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HttpResponse(content, content_type='text/plain')
    except Exception as e:
        return HttpResponse(f"Error reading file: {str(e)}", status=500)

def process_nse_csv(file_path):
    try:
        df = pd.read_csv(file_path, encoding='utf-8', skipinitialspace=True)
        df.columns = df.columns.str.strip().str.upper()
        if 'SYMBOL' not in df.columns:
            raise ValueError("CSV file missing 'SYMBOL' column")
        created_count = 0
        updated_count = 0
        skipped_count = 0
        for index, row in df.iterrows():
            symbol = row['SYMBOL']
            if not symbol or pd.isna(symbol):
                skipped_count += 1
                continue
            price = row.get('LTP', row.get('PREV_CLOSE', row.get('PREV. CLOSE', '0.0')))
            try:
                price = str(price).replace(',', '')
                price_decimal = Decimal(price)
            except (InvalidOperation, ValueError, TypeError) as e:
                skipped_count += 1
                continue
            stock, created = Stock.objects.get_or_create(
                symbol=symbol,
                defaults={
                    'name': row.get('COMPANY_NAME', row.get('SERIES', symbol)),
                    'current_price': price_decimal,
                    'sector': row.get('INDUSTRY', row.get('SECTOR', None)),
                    'exchange': 'NSE'
                }
            )
            if created:
                created_count += 1
            else:
                stock.name = row.get('COMPANY_NAME', row.get('SERIES', stock.name))
                stock.current_price = price_decimal
                stock.sector = row.get('INDUSTRY', row.get('SECTOR', stock.sector))
                stock.exchange = 'NSE'
                stock.save()
                updated_count += 1
        return created_count, updated_count, skipped_count
    except Exception as e:
        print(f"Error processing CSV: {str(e)}")
        raise Exception(f"Error processing CSV: {str(e)}")
